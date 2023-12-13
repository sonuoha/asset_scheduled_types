import openpyxl
from pathlib import Path
import os
import csv


path = Path('./data')

ir = r'data\\CYP_DE Asset Tagging requirement.xlsx'
amr = r'data\\TAS-CYP-CBS-ZWD-REG-XIM-NAP-X0001_C.xlsx'
csv_data = r'data\\scheduled.csv'

amr_data = openpyxl.load_workbook(amr)

ir_data = openpyxl.load_workbook(ir)

amr_validation_ws = amr_data.get_sheet_by_name('Validation Tables')
ir_mep_data = ir_data.get_sheet_by_name('MEP Data Requirement')



ir_types = []
raw_amr_codes = []
amr_codes  = []
amr_code_source = []

scheduled = []

for mep in ir_mep_data.iter_cols(min_row=3, max_col=1):
    for mep_code in mep:
        ir_types.append(mep_code.value)
    
        


for  amr in amr_validation_ws.iter_cols(min_row=844,max_col=1):
    for code in amr:
        raw_amr_codes.append(code.value)

for  amr in amr_validation_ws.iter_cols(min_row=844,min_col=3,max_col=3):
    for code in amr:
        amr_code_source.append(code.value)


for idx, source in enumerate(amr_code_source):
    if source == 'Asset Hive':
        amr_codes.append(raw_amr_codes[idx])

for code in ir_types:
    if code in amr_codes:
        scheduled.append('Yes')
    else:
        scheduled.append('No')

file = open(csv_data, 'w+', newline ='')
 
# writing the data into the file
with file:   
    write = csv.writer(file)
    write.writerows(scheduled)


print(scheduled)


    