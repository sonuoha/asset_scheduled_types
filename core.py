#from update_active_type_code import update_ir_bool_matrix
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
from itertools import islice
import re





twin_codes = r"C:\Users\SamuelOnuoha\John Holland Group\CYP-Digital Engineering - Working Folder\dev\data\twin_codes.xlsx"
ir_file = r"C:\Users\SamuelOnuoha\John Holland Group\CYP-Digital Engineering - Working Folder\dev\data\CYP_DE Asset Tagging requirement.xlsx"
output_path = r"C:\Users\SamuelOnuoha\John Holland Group\CYP-Digital Engineering - Working Folder\dev\data\outputs"

twin_sheetname = 'Export'


ir_sheets = ['MEP Data Requirement', 'RSHP ARC Data Requirement', 'HWW AUD Data Requirement', 'HWW ARC Data Requirement']

#Generating combined twin codes"

twin_code_df = pd.read_excel(twin_codes,'Export')
active_codes = twin_code_df.loc[twin_code_df['Type Description'].notnull()] 
active_codes_list = active_codes.loc[:, 'MM Type'].to_list()

"""
Processind AMR for defined codes
"""
amr_codes = [] #Define amr code list

data_dir = Path(ir_file).parent # return data directory

"""
Iterate over data directory and process amr registers for codes
"""
for dir in data_dir.iterdir():
    if dir.is_file() and 'REG-XIM-NAP' in dir.name:
        wb = load_workbook(dir)
        
        try:
            ws = wb['Validation Tables']
            amr_codes_table = ws.tables['Type_Aux_Lookup']
            print(amr_codes_table.tableColumns[0].LocalName)
            am_cod_ref = amr_codes_table.ref
            pattern = '[0-9]{1,}'
            start_row = int(re.findall(pattern, am_cod_ref)[0])
            end_row = int(re.findall(pattern, am_cod_ref)[1])

            """
            Iterate over excel worksheet using table reference form above
            """
            for row in ws.iter_rows(min_row=start_row, max_row=end_row, max_col=38):
                if row[-1].value != 'Dummy':
                    if len(row[0].value) == 4:
                        #print(row[0].value[:3])
                        amr_codes.append(row[0].value[:3])
                    elif len(row[0].value) > 4:
                        amr_codes.append(row[0].value[-3:])
                        #print(row[0].value[-3:])

                #amr_codes.append(row[0].value)  
                #print(row[0].value, row[-1].value)
            
        except:
            continue       

print(len(amr_codes))
amr_codes = list(dict.fromkeys(amr_codes))
print(amr_codes.count('TPP'))
print(len(amr_codes))





# Active in model function definition
def active_in_model(twin_code_list, type_code_col, active_in_model_col):
    if type_code_col in twin_code_list:
        active_in_model_col = 'Yes'
    else:
        active_in_model_col = 'No'
    return active_in_model_col

# Active in AMR function definition
def scheduled_in_amr(amr_code_list, type_code_col, scheduled_in_amr_col):
    if type_code_col in amr_code_list:
        scheduled_in_amr_col = 'Yes'
    else:
        scheduled_in_amr_col = 'No'
    return scheduled_in_amr_col

# LOI Category function definition
def loi_cat(loi_cat_col, loi2_col, scheduled_in_amr_col):
    if scheduled_in_amr_col == 'Yes' and loi2_col == 'No':
        loi_cat_col = 'LOI 0'
    elif scheduled_in_amr_col == 'No' and loi2_col == 'No':
        loi_cat_col = 'LOI 1'
    elif scheduled_in_amr_col == 'Yes' and loi2_col == 'Yes':
        loi_cat_col = 'LOI 2'
    elif scheduled_in_amr_col == 'No' and loi2_col == 'Yes':
        loi_cat_col = 'LOI 3'
    return loi_cat_col

"""
IR processing
"""

for i in range(0, len(ir_sheets)):
    df = pd.read_excel(ir_file, ir_sheets[i], header=2)
    try:
        df['Active in Model'] = df.apply(lambda row: active_in_model(active_codes_list, row['Code (MM_Type)'], row['Active in Model']), axis=1)
        df['Scheduled in AMR'] = df.apply(lambda row: scheduled_in_amr(amr_codes, row['Code (MM_Type)'], row['Scheduled in AMR']), axis=1)
        #df['LOI Category'] = df.apply(lambda row: "LOI 0" if row['Scheduled in AMR'] == 'Yes' and row['Asset Tag in Model (LOI 1 and 2)'] == 'No'  else 'LOI 1', axis=1)
        df['LOI Category'] = df.apply(lambda row: loi_cat(row['LOI Category'], row['Asset Tag in Model (LOI 1 and 2)'], row['Scheduled in AMR']), axis=1)

    except KeyError:
        #print(df.columns)
        df['Active in Model'] = df.apply(lambda row: active_in_model(active_codes_list, row['Asset Type (MM_Type)'], row['Active in Model']), axis=1)
        df['Scheduled in AMR'] = df.apply(lambda row: scheduled_in_amr(amr_codes, row['Asset Type (MM_Type)'], row['Scheduled in AMR']), axis=1)
        #df['LOI Category'] = df.apply(lambda row: "LOI 0" if row['Scheduled in AMR'] == 'Yes' and row['Asset Tag in Model (LOI 1 and 2)'] == 'No'  else 'LOI 1', axis=1)
        df['LOI Category'] = df.apply(lambda row: loi_cat(row['LOI Category'], row['Asset Tag in Model (LOI 1 and 2)'], row['Scheduled in AMR']), axis=1)
    
    df.to_excel(Path(output_path, ir_sheets[i]).with_suffix('.xlsx'), ir_sheets[i])



