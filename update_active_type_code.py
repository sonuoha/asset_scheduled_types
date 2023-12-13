from openpyxl import workbook, worksheet, load_workbook
import pandas as pd
import numpy as np
import string
from pathlib import Path

# twin_codes = Path("C:\Users\SamuelOnuoha\John Holland Group\CYP-Digital Engineering - Working Folder\dev\data\Utilised MEP MM_Type codes 290623.xlsx")
test_ir_path = r'C:\Users\SamuelOnuoha\John Holland Group\CYP-Digital Engineering - Working Folder\dev\data\CYP_DE Asset Tagging requirement.xlsx'

ir_path = r"C:\Users\SamuelOnuoha\John Holland Group\CYP-Digital Engineering - Documents\General\08_Projectwide Documents\Data requirements\CYP_DE Asset Tagging requirement.xlsx"

twin_code_df = pd.read_excel(twin_codes,'Export')
active_codes = twin_code_df.loc[twin_code_df['Type Description'].notnull()] 

active_codes_list = active_codes.loc[:, 'MM Type'].to_list()


# Live File Processing
ir_wb = load_workbook(test_ir_path)
ir_ws = ir_wb['MEP Data Requirement']



updated_codes = []
for idx, row in enumerate(ir_ws.iter_rows(max_row=10)):
    type_code_col_index = 1000
    active_in_model_col_index = 1000
    for cell in row:
        # print('in the loop')
        print(idx, cell.coordinate)
        if idx == 0:
            print(type(int(row[0].coordinate[-1])), int(row[0].coordinate[-1]), type(row[0].coordinate[-1]), row[0].coordinate[-1], int(row[0].coordinate[-1]) == 1)

        if str(cell.value ) == 'Asset Type (MM_Type)':
            type_code_col_index = cell.col_idx
            print(cell.col_idx, cell.value, str(cell.value ) == 'Asset Type (MM_Type)')
        elif cell.value == 'Active in Model':
            active_in_model_col_index = cell.col_idx
            print(cell.col_idx, cell.value, cell.value == 'Active in Model')
            print('breaking')
            break

    #print(type_code_col_index, active_in_model_col_index)
    # if row[0].value in active_codes_list:
    #     row[-1].value = 'Yes'
    #     updated_codes.append(row[0].value)
    #     # print(row[0].value)
    # else:
    #     row[-1].value = 'No'

# ir_wb.save(ir_path)

def update_ir_bool_matrix(twin_code_path, twin_code_sheetname: str, ir_path, ir_sheetname: str, type_col_name = 'Asset Type (MM_Type)', bool_col_name = 'Active in Model', min_row=None, max_row=None, min_col=None, max_col=None, values_only=False):

    '''
    This function iterates over a Pandas dataframe of active type codes (exported from the willow twin), creating a temporary list. It then iterates over project 
    Information requirement checking if the a defined project type code has been utilized in the model as the list from Willow Twin export. Where it exist, column
    corrsponding to "Active in Model" is updated to "Yes" and "Noi" otherwise.

    The function expectes the first column to correspond to the Type Code coulumn and last column to the Active in Model column respectively and the first row corresponding to the
    first type code definition record in the dataset (Not the header).

    Function Parameters:

    twin_code_path: path to excel file of asset type codes exported from willow twin
    twin_code_sheetname: Sheet name of the twin code excel file (Type = String)
    ir_path: path to excel file of the project information requirement
    ir_sheetname: Sheet name of the Information Requirement file (Type = String)
    min_row: 1 based row number of the information requirement file record
    max_row: 1 based row  number of the last row of the Information Requirement to process
    min_col: Min Information Requirement File column to start iteration. Should correspond to the Type Code column with Active in Model to the right.
    max_col: Max Information Requirement File column to end iteration. Should correspond to the "Active in Model" column with Active in Model to the right.
    values_only: whether only cell values should be returned

    '''
    twin_code_df = pd.read_excel(twin_code_path, twin_code_sheetname)
    active_codes_list = twin_code_df.loc[:, 'MM Type'].to_list()

    ir_wb = load_workbook(filename=ir_path)
    ir_ws = ir_wb[ir_sheetname]

    type_code_col_index = 0
    active_in_model_col_index = 0

    updated_codes = []
    for idx, row in enumerate(ir_ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=values_only)):
      
        for cell in row:
            if min_row == None:
                if str(cell.value ) == type_col_name:
                    type_code_col_index = cell.col_idx      # Retrieve Type Column Excel Index
                    print(cell.col_idx )
            else:
                type_code_col_index = min_row

            if max_col == None:
                if cell.value == bool_col_name:
                    active_in_model_col_index = cell.col_idx
                    break
            else:
                active_in_model_col_index = max_col
                break
        
        print(int(row[(int(type_code_col_index)- 1)].coordinate[-1]))
        if idx == 0 & int(row[(int(type_code_col_index)- 1)].coordinate[-1]) == 1:
            print(int(row[(int(type_code_col_index)- 1)].coordinate[-1]) == 1)
            continue
        else:
            if row[(int(type_code_col_index)- 1)].value in active_codes_list:
                row[-1].value = 'Yes'
                updated_codes.append(row[type_code_col_index-1].value)
                print('Updated Yes')
            else:
                row[(int(active_in_model_col_index) - 1)].value = 'No'
                print('Updated No')
        
    ir_wb.save(ir_path)
    print('The following type codes were updated:', updated_codes)
    return


 